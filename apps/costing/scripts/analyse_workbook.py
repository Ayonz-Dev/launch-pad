#!/usr/bin/env python3
"""Extract the visible structure and formulas from the source costing workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def serialise(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()

    formula_book = load_workbook(args.workbook, data_only=False, read_only=False)
    value_book = load_workbook(args.workbook, data_only=True, read_only=False)
    result: dict[str, Any] = {
        "file": str(args.workbook),
        "sheets": [],
        "defined_names": [
            {"name": item.name, "attr_text": item.attr_text}
            for item in formula_book.defined_names.values()
        ],
    }

    for sheet in formula_book.worksheets:
        value_sheet = value_book[sheet.title]
        cells = []
        formulas = []
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                entry = {
                    "cell": cell.coordinate,
                    "value": serialise(cell.value),
                    "cached": serialise(value_sheet[cell.coordinate].value),
                    "number_format": cell.number_format,
                    "fill": cell.fill.fgColor.rgb or cell.fill.fgColor.indexed,
                    "font_color": serialise(
                        cell.font.color.type + ":" + str(cell.font.color.rgb or cell.font.color.indexed)
                        if cell.font.color
                        else None
                    ),
                    "locked": cell.protection.locked,
                    "hidden": cell.protection.hidden,
                }
                cells.append(entry)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formulas.append(entry)

        result["sheets"].append(
            {
                "name": sheet.title,
                "state": sheet.sheet_state,
                "dimensions": sheet.calculate_dimension(),
                "freeze_panes": serialise(sheet.freeze_panes),
                "merged_ranges": [str(item) for item in sheet.merged_cells.ranges],
                "cells": cells,
                "formula_count": len(formulas),
                "formulas": formulas,
            }
        )

    print(json.dumps(result, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
